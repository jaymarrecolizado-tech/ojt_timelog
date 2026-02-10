from datetime import date, timedelta, datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from calendar import monthrange
from ...core.database import get_db
from ...models.models import User, Student, TimeLog, Holiday
from ...services.pdf_service import dtr_generator
from ...utils.time_calc import calculate_hours_for_day, get_day_status
from .auth import get_current_user, require_role

router = APIRouter()


def calculate_hours_for_day(logs_for_date: list) -> tuple[float, dict]:
    """Calculate hours for a single day based on time logs."""
    day_data = {"am_in": None, "am_out": None, "pm_in": None, "pm_out": None}
    am_in_time = None
    am_out_time = None
    pm_in_time = None
    pm_out_time = None

    for log in logs_for_date:
        time_str = log.timestamp.strftime("%I:%M").lstrip("0")
        if log.log_category == "AM":
            if log.log_type == "IN":
                day_data["am_in"] = time_str
                am_in_time = log.timestamp
            else:
                day_data["am_out"] = time_str
                am_out_time = log.timestamp
        else:
            if log.log_type == "IN":
                day_data["pm_in"] = time_str
                pm_in_time = log.timestamp
            else:
                day_data["pm_out"] = time_str
                pm_out_time = log.timestamp

    hours = 0.0
    if am_in_time and am_out_time:
        diff = (am_out_time - am_in_time).total_seconds() / 3600
        hours += max(0, min(diff, 4))
    if pm_in_time and pm_out_time:
        diff = (pm_out_time - pm_in_time).total_seconds() / 3600
        hours += max(0, min(diff, 4))

    return round(hours, 2), day_data


@router.get("/dtr", response_model=dict)
async def get_dtr(
    student_id: str | None = None,
    month: int = Query(...),
    year: int = Query(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if user.role == "student":
        result = await db.execute(select(Student).where(Student.user_id == user.id))
        student = result.scalar_one_or_none()
        if not student:
            return {"success": False, "error": "Student not found"}
        target_student_id = student.id
        student_record = student
    else:
        if not student_id:
            return {"success": False, "error": "student_id required"}
        result = await db.execute(select(Student).where(Student.id == student_id))
        student_record = result.scalar_one_or_none()
        target_student_id = student_id

    if not student_record:
        return {"success": False, "error": "Student not found"}

    _, days_in_month = monthrange(year, month)
    from_date = date(year, month, 1)
    to_date = date(year, month, days_in_month)

    result = await db.execute(
        select(TimeLog)
        .where(
            TimeLog.student_id == target_student_id,
            TimeLog.date >= from_date,
            TimeLog.date <= to_date,
        )
        .order_by(TimeLog.date, TimeLog.timestamp)
    )
    logs = list(result.scalars().all())

    result = await db.execute(select(Holiday))
    holidays = list(result.scalars().all())

    result = await db.execute(
        select(TimeLog)
        .where(TimeLog.student_id == target_student_id)
        .order_by(TimeLog.date, TimeLog.timestamp)
    )
    all_logs = list(result.scalars().all())

    logs_by_date = {}
    for log in logs:
        if log.date not in logs_by_date:
            logs_by_date[log.date] = []
        logs_by_date[log.date].append(log)

    all_logs_by_date = {}
    for log in all_logs:
        if log.date not in all_logs_by_date:
            all_logs_by_date[log.date] = []
        all_logs_by_date[log.date].append(log)

    rows = []
    monthly_hours = 0.0
    current = from_date

    while current <= to_date:
        day_name = current.strftime("%A")
        day_logs = logs_by_date.get(current, [])

        if day_logs:
            hours, day_data = calculate_hours_for_day(day_logs)
            if hours >= 7.5:
                remarks = ""
            else:
                remarks = "INCOMPLETE"
        else:
            hours = 0.0
            day_data = {"am_in": "", "am_out": "", "pm_in": "", "pm_out": ""}
            remarks = get_day_status(day_data, current, holidays)

        hours_str = f"{hours:g} hrs" if hours > 0 else ""
        monthly_hours += hours

        rows.append(
            {
                "date": current.strftime("%B %d, %Y"),
                "day": day_name,
                "am_in": day_data.get("am_in") or "",
                "am_out": day_data.get("am_out") or "",
                "pm_in": day_data.get("pm_in") or "",
                "pm_out": day_data.get("pm_out") or "",
                "hours_rendered": hours_str,
                "remarks": remarks,
            }
        )
        current += timedelta(days=1)

    accumulated_hours = 0.0
    for day_logs in all_logs_by_date.values():
        hours, _ = calculate_hours_for_day(day_logs)
        accumulated_hours += hours

    middle_initial = (
        f" {student_record.middle_name[0]}." if student_record.middle_name else ""
    )
    full_name = f"{student_record.last_name.upper()}, {student_record.first_name}{middle_initial}"
    ojt_period = ""
    if student_record.ojt_start and student_record.ojt_end:
        ojt_period = f"{student_record.ojt_start.strftime('%B %d, %Y')} - {student_record.ojt_end.strftime('%B %d, %Y')}"

    return {
        "success": True,
        "data": {
            "student": {
                "full_name": full_name,
                "student_id_no": student_record.student_id_no,
                "department": student_record.department,
                "program": student_record.program,
                "company": student_record.company or "",
                "ojt_period": ojt_period,
            },
            "period": f"{date(year, month, 1).strftime('%B')} {year}",
            "rows": rows,
            "totals": {
                "monthly_hours": round(monthly_hours, 1),
                "accumulated_hours": round(accumulated_hours, 1),
                "required_hours": float(student_record.required_hours),
                "remaining_hours": round(
                    float(student_record.required_hours) - accumulated_hours, 1
                ),
                "completion_percentage": round(
                    (accumulated_hours / float(student_record.required_hours)) * 100, 1
                )
                if student_record.required_hours
                else 0,
            },
        },
    }


from io import BytesIO


@router.get("/dtr/pdf")
async def get_summary(
    month: int = Query(...),
    year: int = Query(...),
    department: str = "all",
    user: User = Depends(require_role(["admin", "super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Student))
    students = result.scalars().all()

    student_list = []
    for s in students:
        if department != "all" and s.department != department:
            continue

        result = await db.execute(
            select(TimeLog)
            .where(
                TimeLog.student_id == s.id,
                TimeLog.date >= date(year, month, 1),
                TimeLog.date <= date(year, month, 28),
            )
            .order_by(TimeLog.date, TimeLog.timestamp)
        )
        logs = result.scalars().all()

        logs_by_date = {}
        for log in logs:
            if log.date not in logs_by_date:
                logs_by_date[log.date] = []
            logs_by_date[log.date].append(log)

        monthly_hours = 0.0
        for day_logs in logs_by_date.values():
            hours, _ = calculate_hours_for_day(day_logs)
            monthly_hours += hours

        student_list.append(
            {
                "student_id_no": s.student_id_no,
                "name": f"{s.last_name}, {s.first_name}",
                "department": s.department,
                "monthly_hours": round(monthly_hours, 1),
                "accumulated_hours": round(monthly_hours, 1),
                "required_hours": float(s.required_hours),
                "completion": f"{round((monthly_hours / float(s.required_hours)) * 100, 1)}%"
                if s.required_hours
                else "0%",
                "days_present": len(logs_by_date),
                "days_absent": 0,
                "status": "ON_TRACK" if monthly_hours >= 80 else "BEHIND",
            }
        )

    return {
        "success": True,
        "data": {
            "period": f"{date(year, month, 1).strftime('%B')} {year}",
            "students": student_list,
        },
    }


@router.get("/dtr/pdf")
async def get_dtr_pdf(
    student_id: str | None = None,
    student_id_no: str | None = None,
    month: int = Query(...),
    year: int = Query(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    target_student_id = student_id

    if student_id_no and not target_student_id:
        result = await db.execute(
            select(Student).where(Student.student_id_no == student_id_no)
        )
        student = result.scalar_one_or_none()
        if student:
            target_student_id = str(student.id)

    if not target_student_id:
        if user.role == "student":
            result = await db.execute(select(Student).where(Student.user_id == user.id))
            student = result.scalar_one_or_none()
            if student:
                target_student_id = str(student.id)

        if not target_student_id:
            return Response(content="Student not found", status_code=404)

    dtr_response = await get_dtr(target_student_id, month, year, user, db)

    if not dtr_response.get("success"):
        return Response(
            content=dtr_response.get("error", "Error generating DTR"), status_code=400
        )

    dtr_data = dtr_response["data"]

    pdf_bytes = dtr_generator.generate_dtr_pdf(
        student_info=dtr_data["student"],
        period=dtr_data["period"],
        rows=dtr_data["rows"],
        totals=dtr_data["totals"],
    )

    filename = f"DTR_{dtr_data['student']['student_id_no']}_{month}_{year}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dtr/excel")
async def get_dtr_excel(
    student_id: str | None = None,
    month: int = Query(...),
    year: int = Query(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    dtr_response = await get_dtr(student_id, month, year, user, db)

    if not dtr_response.get("success"):
        return Response(
            content=dtr_response.get("error", "Error generating DTR"), status_code=400
        )

    dtr_data = dtr_response["data"]

    wb = Workbook()
    ws = wb.active
    ws.title = "DTR"

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "DAILY TIME RECORD"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Name:"
    ws["B3"] = dtr_data["student"]["full_name"]
    ws["A4"] = "Student ID:"
    ws["B4"] = dtr_data["student"]["student_id_no"]
    ws["A5"] = "Department:"
    ws["B5"] = dtr_data["student"]["department"]
    ws["A6"] = "Company:"
    ws["B6"] = dtr_data["student"]["company"]
    ws["A7"] = "Period:"
    ws["B7"] = dtr_data["period"]

    headers = ["Date", "Day", "AM In", "AM Out", "PM In", "PM Out", "Hours", "Remarks"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(dtr_data["rows"], 10):
        ws.cell(row=row_idx, column=1, value=row["date"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=row["day"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=row["am_in"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=row["am_out"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=row["pm_in"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=row["pm_out"]).border = thin_border
        ws.cell(row=row_idx, column=7, value=row["hours_rendered"]).border = thin_border
        ws.cell(row=row_idx, column=8, value=row["remarks"]).border = thin_border

    totals = dtr_data["totals"]
    total_row = len(dtr_data["rows"]) + 11
    ws.cell(row=total_row, column=1, value="TOTALS").font = header_font
    ws.cell(
        row=total_row, column=7, value=f"{totals['monthly_hours']} hrs"
    ).font = header_font

    ws.cell(
        row=total_row + 2,
        column=1,
        value=f"Accumulated: {totals['accumulated_hours']} hrs",
    )
    ws.cell(
        row=total_row + 3, column=1, value=f"Required: {totals['required_hours']} hrs"
    )
    ws.cell(
        row=total_row + 4, column=1, value=f"Remaining: {totals['remaining_hours']} hrs"
    )
    ws.cell(
        row=total_row + 5,
        column=1,
        value=f"Completion: {totals['completion_percentage']}%",
    )

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DTR_{dtr_data['student']['student_id_no']}_{month}_{year}.xlsx"

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary/excel")
async def get_summary_excel(
    month: int = Query(...),
    year: int = Query(...),
    department: str = "all",
    user: User = Depends(require_role(["admin", "super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    summary_response = await get_summary(month, year, department, user, db)

    if not summary_response.get("success"):
        return Response(content="Error generating summary", status_code=400)

    data = summary_response["data"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"] = "OJT ATTENDANCE SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Period: {data['period']}"

    headers = [
        "Student ID",
        "Name",
        "Department",
        "Monthly Hours",
        "Accumulated",
        "Required",
        "Completion",
        "Days Present",
        "Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, student in enumerate(data["students"], 5):
        ws.cell(
            row=row_idx, column=1, value=student["student_id_no"]
        ).border = thin_border
        ws.cell(row=row_idx, column=2, value=student["name"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=student["department"]).border = thin_border
        ws.cell(
            row=row_idx, column=4, value=student["monthly_hours"]
        ).border = thin_border
        ws.cell(
            row=row_idx, column=5, value=student["accumulated_hours"]
        ).border = thin_border
        ws.cell(
            row=row_idx, column=6, value=student["required_hours"]
        ).border = thin_border
        ws.cell(row=row_idx, column=7, value=student["completion"]).border = thin_border
        ws.cell(
            row=row_idx, column=8, value=student["days_present"]
        ).border = thin_border
        ws.cell(row=row_idx, column=9, value=student["status"]).border = thin_border

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"OJT_Summary_{month}_{year}.xlsx"

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/batch/dtr/pdf")
async def get_batch_dtr_pdf(
    month: int = Query(...),
    year: int = Query(...),
    department: str = "all",
    user: User = Depends(require_role(["admin", "super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    import zipfile
    import io

    query = select(Student)
    if department != "all":
        query = query.where(Student.department == department)

    result = await db.execute(query)
    students = result.scalars().all()

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for student in students:
            dtr_response = await get_dtr(str(student.id), month, year, user, db)

            if dtr_response.get("success"):
                dtr_data = dtr_response["data"]
                pdf_bytes = dtr_generator.generate_dtr_pdf(
                    student_info=dtr_data["student"],
                    period=dtr_data["period"],
                    rows=dtr_data["rows"],
                    totals=dtr_data["totals"],
                )

                filename = f"DTR_{student.student_id_no}_{month}_{year}.pdf"
                zipf.writestr(filename, pdf_bytes)

    zip_buffer.seek(0)

    return Response(
        content=zip_buffer.read(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="DTR_Batch_{month}_{year}.zip"'
        },
    )
